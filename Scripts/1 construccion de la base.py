# -*- coding: utf-8 -*-
"""
Created on Mon Mar 28 17:36:02 2022

@author: hugoferq y carlosRV0803
"""

#Importar las librerias
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment  
pd.options.mode.chained_assignment = None

#Configuro la direccion de trabajo
work =  Path(r'D:\Trabajo\AITeacherAllocation')

#Cargo las bases de insumos
racio_2019 = pd.read_csv(work/r'Raw Data\Base 2019.csv')

racio_2020 = pd.read_stata(work/r'Raw Data\\Racio 2020.dta')
racio_2021 = pd.read_stata(work/r'Raw Data\\Racio 2021.dta')
padron_gg1 = pd.read_stata(work/r'Raw Data\\Padron GG1.dta')
#Renombro alguans variables
racio_2021.rename(columns={'numeroseccion':'secciones_necesarias'}, inplace=True)


#Funcion para limpiar y homogenizar las bases de datos
def cargar_base(df,anio):
    """
    Esta funcion limpia y homogeniza los resultados de racionalizacion

    Parameters
    ----------
    df :   TYPE: DataFrame para limpiar y homogenizar
           DataFrame para limpiar y homogenizar.
    anio : TYPE. Int
           Año de la base de datos

    Returns
    -------
    df_ok : TYPE. DataFrame
            DataFrame limpio y homogenizada

    """
    #Variables a usar
    all_columns = df.columns.tolist()
    
    #Datos de identificacion (salen del padron gg1)
    identificacion_padron = ['cod_mod','niv_mod', 'd_niv_mod','gestion','d_gestion','ges_dep','d_ges_dep','ubigeo',
                         'd_dpto','d_prov','d_dist','d_region','codooii','d_dreugel','nlat_ie','nlong_ie',
                         'estado','d_estado','region','tipo_entidad',f'jec_{anio}'] 
    asignaciones_temporales=[f'rural_upp_{anio}',f'vraem_upp_{anio}',f'fron_upp_{anio}',f'bilin_upp_{anio}',f'tipie_upp_{anio}']
    identificacion = ['cod_mod','codlocal']
    padron_gg1_short = padron_gg1.loc[padron_gg1['anexo']=='0' ,identificacion_padron+asignaciones_temporales]
    padron_gg1_short.rename(columns={f'rural_upp_{anio}':'ruralidad',
                                     f'vraem_upp_{anio}':'vraem',
                                     f'fron_upp_{anio}':'frontera',
                                     f'bilin_upp_{anio}':'bilingue',
                                     f'tipie_upp_{anio}':'caracteristica',
                                     f'jec_{anio}':'jec'},inplace=True)
    
    #PEA evaluada
    for cargo in ['dir','sub_dir']:
        df[f'{cargo}_n'] = df[f'{cargo}_des_org']+df[f'{cargo}_des_ev']
        df[f'{cargo}_c'] = df[f'{cargo}_des_ev']+df[f'{cargo}_enc_ev']+df[f'{cargo}_vac_ev']
	
    for cargo in ['jer','doc','otro_doc','aux']:
        df[f'{cargo}_n']=df[f'{cargo}_nom_org']+df[f'{cargo}_nom_ev']
        df[f'{cargo}_c']=df[f'{cargo}_con_org']+df[f'{cargo}_con_ev']+df[f'{cargo}_vac_org']+df[f'{cargo}_vac_ev']

    pea_evaluada = []
    for cargo in ['dir','sub_dir','jer','doc','otro_doc','aux']:
        for sit in ['n','c']:
            pea_evaluada.append(f'{cargo}_{sit}')    
    
    #Matricula     
        # Cambiando el nombre de la matricula para que todas los anios tengan la misma extension
    df = df.rename(columns={f'cant0_alum_{anio}':'cant0 (t)',
                            f'cant1_alum_{anio}': 'cant1 (t)', 
                            f'cant2_alum_{anio}': 'cant2 (t)', 
                            f'cant3_alum_{anio}': 'cant3 (t)', 
                            f'cant4_alum_{anio}': 'cant4 (t)', 
                            f'cant5_alum_{anio}': 'cant5 (t)',
                            f'cant6_alum_{anio}': 'cant6 (t)'})

    df = df.rename(columns={f'cant0_inclusivo_{anio}': 'inclu0 (t)',
                            f'cant1_inclusivo_{anio}': 'inclu1 (t)', 
                            f'cant2_inclusivo_{anio}': 'inclu2 (t)', 
                            f'cant3_inclusivo_{anio}': 'inclu3 (t)', 
                            f'cant4_inclusivo_{anio}': 'inclu4 (t)', 
                            f'cant5_inclusivo_{anio}': 'inclu5 (t)',
                            f'cant6_inclusivo_{anio}': 'inclu6 (t)'})

    lag = 0
    anios = [anio - i for i in [1,2,3,4]]
    for lapso in anios:
      lag = lag + 1
      df = df.rename(columns={f'cant0_alum_{lapso}': f'cant0 (t-{lag})', 
                              f'cant1_alum_{lapso}': f'cant1 (t-{lag})',
                              f'cant2_alum_{lapso}': f'cant2 (t-{lag})', 
                              f'cant3_alum_{lapso}': f'cant3 (t-{lag})',
                              f'cant4_alum_{lapso}': f'cant4 (t-{lag})', 
                              f'cant5_alum_{lapso}': f'cant5 (t-{lag})', 
                              f'cant6_alum_{lapso}': f'cant6 (t-{lag})'})
      
      df = df.rename(columns={f'cant0_inclusivo_{lapso}': f'inclu0 (t-{lag})', 
                              f'cant1_inclusivo_{lapso}': f'inclu1 (t-{lag})',
                              f'cant2_inclusivo_{lapso}': f'inclu2 (t-{lag})', 
                              f'cant3_inclusivo_{lapso}': f'inclu3 (t-{lag})',
                              f'cant4_inclusivo_{lapso}': f'inclu4 (t-{lag})', 
                              f'cant5_inclusivo_{lapso}': f'inclu5 (t-{lag})', 
                              f'cant6_inclusivo_{lapso}': f'inclu6 (t-{lag})'})
    
    matricula_rename = [x for x in df.columns.to_list() if x.find(' (t')!=-1]
    
    #Datos de la evaluacion
    datos_evaluacion = ['usuario_minedu','bolsa_nexus','bolsa_sira','secciones_necesarias']
    #Resultados
    requerimientos = [x for x in all_columns if x.startswith('req') and not x.find('req_exd')!=-1]
    excedentes = [x for x in all_columns if x.find('exd')!=-1 and x.endswith(f'{anio}') and not x.find('tot_')!=-1 ]
    #Agrego datos del padron
    racio_short = df[identificacion+pea_evaluada+matricula_rename+datos_evaluacion+requerimientos+excedentes] 
    #Base consolidada   
    df_ok = pd.merge(racio_short,padron_gg1_short,on=['cod_mod'],how='left',validate='1:1')
    #Homogenizo las variables
    for col in df_ok.columns:
        df_ok.rename(columns={col:col.replace(f'_{anio}','')},inplace=True)
        
    df_ok = df_ok.rename(columns={'doc_n':'doc_aula_n',
                                  'doc_c':'doc_aula_c',
                                  'doc_no_aula_exd':'otro_doc_exd',
                                  'director_exd':'dir_exd',
                                  'subdirector_exd':'sub_dir_exd',
                                  'jerarquico_exd':'jer_exd',
                                  'auxiliar_exd':'aux_exd',
                                  'req_dir':'dir_req',
                                  'req_sub':'sub_dir_req',
                                  'req_jer':'jer_req',
                                  'req_prof':'doc_aula_req',
                                  'req_aux':'aux_req',
                                  'req_fisica':'fisica_req',
                                  'req_aip':'aip_req'})
    df_ok['year'] = anio
    return df_ok




#Consolido la base
racio_2020_ok = cargar_base(racio_2020,2020)
racio_2021_ok = cargar_base(racio_2021,2021)
df = racio_2020_ok.append(racio_2021_ok)
df = df.append(racio_2019)
#Exporto la base
df.to_csv(work/r'Results\Base consolidada.csv')
df.to_csv('D:\OneDrive\Trabajo\Minedu\AI teacher allocation data\Results\Base consolidada.csv')

    


# Build a data dictionary
all_columns = df.columns.to_list()
    # cargos
cargos=[x for x in df.columns.to_list() if x.find('doc_aula')!=-1 or 
                                            x.find('otro_doc')!=-1 or
                                            x.find('dir')!=-1 or
                                            x.find('sub_dir')!=-1 or
                                            x.find('jer')!=-1 or
                                            x.find('aip')!=-1 or
                                            x.find('fisica')!=-1 or
                                            x.find('aux')!=-1]
pea = ['']*len(cargos)
df_dd_cargos = pd.DataFrame(list(zip(cargos, pea)),columns =['Variable', 'Descripcion'])                                          
dict_cargos={'doc_aula':'docente de aula','otro_doc':'otro docente', 
             'dir':'director','sub_dir':'subdirector','jer':'jerarquico',
             'aux':'auxiliar','aip':'docente AIP','fisica':'docente de educacion fisica'}
dict_contrato={'_n':'nombrado', '_c':'contratado'}
dict_estado={'_exd':'excedente','_req':'requerimiento'}
for k,v in dict_cargos.items():
    df_dd_cargos.loc[df_dd_cargos['Variable'].str.find(f'{k}')!=-1,'cargos']=f'{v}'
for k,v in dict_contrato.items():
    df_dd_cargos.loc[df_dd_cargos['Variable'].str.find(f'{k}')!=-1,'contrato']=f'{v}'
for k,v in dict_estado.items():
    df_dd_cargos.loc[df_dd_cargos['Variable'].str.find(f'{k}')!=-1,'estado']=f'- {v}'    
for x in ['cargos','contrato','estado']:
    df_dd_cargos.loc[df_dd_cargos[f'{x}'].str.find('NaN')!=-1,f'{x}']= ''
df_dd_cargos['Descripcion']=df_dd_cargos['cargos']+' '+df_dd_cargos['contrato'] +df_dd_cargos['estado']  
df_dd_cargos['Descripcion']
df_dd_cargos=df_dd_cargos[['Variable','Descripcion']]

    # otras variables
dict_otros={'cod_mod':'codigo modular (identificaador del servicio educativo)',
'codlocal':'codigo local (identificaador del local)',
'usuario_minedu':'usuario que realizo la evaluacion',
'bolsa_nexus':'bolsa de horas registrada en nexus',
'bolsa_sira':'bolsa de horas registrada en sira (resultado del proceso)',
'secciones_necesarias':'secciones necesarias (resultado del proceso)',
'niv_mod':'codigo de nivel',
'd_niv_mod':'descripcion del nivel',
'gestion':'codigo de gestion',
'd_gestion':'descripcion de gestion',
'ges_dep':'codigo de dependencia',
'd_ges_dep':'descripcion de la dependencia',
'ubigeo':'ubigeo',
'd_dpto':'departamento',
'd_prov':'provincia',
'd_dist':'distrito',
'd_region':'direccion regional de educacion (DRE)',
'codooii':'codigo de ugel (identificador de ugel)',
'd_dreugel':'descripcion de la ugel',
'nlat_ie':'latitud del local educativo',
'nlong_ie':'longitud del local educativo',
'estado':'codigo del estado del servicio educativo',
'd_estado':'descripcion del estado del servicio educativo',
'region':'region',
'tipo_entidad':'entidades ejecutoras u operativas',
'jec':'servicio educativo con jornada escolar completa(jec)',
'ruralidad':'gradiente de ruralidad del servicio educativo',
'vraem':'servicio educativo localizado en el valle del rio ene-mantaro',
'frontera':'servicio educativo ubicado en la frontera',
'bilingue':'servicio educativo con educacion bilingue',
'caracteristica':'caracteristica del servicio educativo (unidocente,multigrado o polidocente)',
'year':'año de evaluacion'}
df_dd_otros=pd.DataFrame(list(zip(dict_otros.keys(),dict_otros.values())),columns =['Variable', 'Descripcion'])
df_dd_descripcion = df_dd_otros.append(df_dd_cargos)
def make_my_data_dictionary(dataFrame):
    """
    Parameters
    ----------
    dataFrame : TYPE: DataFrame
        El DataFrame que genera el diccionario de datos

    Returns
    -------
    df_DD : TYPE: DataFrame
        Diccionario de datos
    """
    col_ = list(df_dd_descripcion['Variable'])
    df_DataDict = {}

    for col in col_:
            df_DataDict[col] = {
                           'Tipo': str(dataFrame.dtypes[col]),
                           'Longitud': len(dataFrame[col]),
                           'Missing': sum(dataFrame[col].isna()),
                           'Tamaño(memoria)': dataFrame.memory_usage()[col],
                            }
    df_DD = pd.DataFrame(df_DataDict).transpose()   
    df_DD.reset_index(inplace=True)
    df_DD.rename(columns={'index':'Variable'},inplace=True)
    return df_DD

df_dd = make_my_data_dictionary(df)
df_dd = pd.merge(left=df_dd,right=df_dd_descripcion)
df_dd.to_excel(work/r'Documents\diccionario de datos.xlsx',sheet_name='principal',index=False)

    #matricula
workbook = load_workbook(filename=work/r'Documents\diccionario de datos.xlsx')
ws = workbook.create_sheet(title='matricula')
for x in ['A','D','G']:
    #Encabezado
    ws[f'{x}1'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'{x}1']='Nivel'
    ws[f'{x}1'].font = Font(bold=True)
    ws.merge_cells(f'{x}2:{x}7')
    ws[f'{x}2'].alignment = Alignment(horizontal='center', vertical='center')    
ws['A2']='Inicial'
ws['D2']='Primaria'
ws['G2']='Secundaria'
for x in ['B','E','H']:
    ws[f'{x}1']='Nomenclatura'
    ws[f'{x}1'].font = Font(bold=True)
    ws.alignment = Alignment(horizontal='center', vertical='center')
for x in range(2,9):
    ws[f'B{x}']=ws[f'E{x}']=ws[f'H{x}']=x-2

for x in ['C','F','I']:
    ws[f'{x}1']='Descripcion'
    ws[f'{x}1'].font = Font(bold=True)
    ws.alignment = Alignment(horizontal='center', vertical='center')
#inicial
ws['C2']='De 0 hasta 12 meses'
ws['C3']='De 12 hasta 24 meses'
ws['C4']='De 24 hasta 36 meses'
ws['C5']='3 años'
ws['C6']='4 años'
ws['C7']='5 años'
ws['C8']='No aplica'
#primaria
ws['F2']='No aplica'
ws['F3']='Primero de primaria'
ws['F4']='Segundo de primaria'
ws['F5']='Tercero de primaria'
ws['F6']='Cuarto de primaria'
ws['F7']='Quinto de primaria'
ws['F8']='Sexto de primaria'
#secundaria
ws['I2']='No aplica'
ws['I3']='Primero de secundaria'
ws['I4']='Segundo de secundaria'
ws['I5']='Tercero de secundaria'
ws['I6']='Cuarto de secundaria'
ws['I7']='Quinto de secundaria'
ws['I8']='No aplica'

ws['A10'] = 'inclu: indica alumnos inclusivos'
ws['A11'] = 'cant: indica alumnos regulares'
workbook.save(work/r'Documents\diccionario de datos.xlsx')